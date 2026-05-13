"""
Tratamento do Odonto da VSP (SAMP) - divisão em SINDSEG e SINDIVIGILANTES.

Cada sindicato tem 4 boletos (8 no total):
  - Odonto - Funcionários (R$ 15,00)
  - Odonto - Dependentes (R$ 15,00)
  - Afetos                (R$ 18,50)
  - Telemedicina          (R$ 10,00)

Identificação:
  1. Pelo nome do arquivo (SINDSEG ou SINDIVIGILANTES).
  2. Pela pasta dentro do ZIP (ODONTO SINDSEG / ODONTO SINDIVIGILANTES).
  3. Pelo valor majoritário (excluindo zerados) caso 1 e 2 falhem.

Tipo de boleto identificado pelo valor majoritário (excluindo zerados):
  - 15,00 -> pode ser Funcionários OU Dependentes.
      Como diferenciar? Pelo arquivo: "Controle de Pagamentos" tem coluna
      "Nome Dependente" -> Odonto Dependentes. Os "Calculo relação" são Funcionários.
  - 18,50 -> Afetos
  - 10,00 -> Telemedicina

Saída: planilha tratada por boleto + uma aba consolidada por CCusto.
"""

import io
import re
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import (
    ler_excel_qualquer,
    limpar_cpf,
    limpar_nome,
    parse_valor_br,
    codigo_nome_quebra,
)


# Mapeamento valor majoritário -> tipo de boleto (quando NÃO é dependente)
VALOR_PARA_TIPO_FUNC = {
    15.00: "Odonto - Funcionários",
    18.50: "Afetos",
    10.00: "Telemedicina",
}

# Tolerância para comparação de valores
TOL = 0.01


def identificar_sindicato(nome_arquivo: str, pasta: str = "") -> str:
    """Retorna 'SINDSEG' ou 'SINDIVIGILANTES' (ou '' se não identificado)."""
    base = f"{nome_arquivo} {pasta}".upper()
    if "SINDIVIGILANTES" in base:
        return "SINDIVIGILANTES"
    if "SINDSEG" in base:
        return "SINDSEG"
    return ""


def identificar_tipo_boleto(df_raw: pd.DataFrame, nome_arquivo: str = "") -> str:
    """
    Identifica o tipo de boleto a partir do conteúdo + nome do arquivo.

    Retorna um dos: 'Odonto - Funcionários', 'Odonto - Dependentes',
                    'Afetos', 'Telemedicina'.
    """
    nome_up = nome_arquivo.upper()

    # Se for "Controle de Pagamentos" -> Odonto Dependentes (tem "Nome Dependente")
    if "CONTROLE DE PAGAMENTOS" in nome_up or "DEPENDENTE" in nome_up:
        return "Odonto - Dependentes"

    # Procurar a coluna "Nome Dependente" no conteúdo (qualquer linha do header)
    for i in range(min(5, len(df_raw))):
        row = [str(x).strip() for x in df_raw.iloc[i].tolist()]
        for cell in row:
            if "Nome Dependente" in cell:
                return "Odonto - Dependentes"

    # Caso contrário, identifica pelo valor majoritário
    valor_maj = _valor_majoritario(df_raw)
    if valor_maj is None:
        return "Odonto - Funcionários"

    if abs(valor_maj - 18.50) < TOL:
        return "Afetos"
    if abs(valor_maj - 10.00) < TOL:
        return "Telemedicina"
    if abs(valor_maj - 15.00) < TOL:
        return "Odonto - Funcionários"
    # Fallback
    return "Odonto - Funcionários"


def _valor_majoritario(df_raw: pd.DataFrame) -> Optional[float]:
    """Retorna o valor que mais aparece na planilha (ignorando 0 e NaN)."""
    valores = []
    # Procurar a coluna "Valor"
    valor_col_idx = None
    header_row = None
    for i in range(min(8, len(df_raw))):
        for j, cell in enumerate(df_raw.iloc[i].tolist()):
            if str(cell).strip().lower() == "valor":
                valor_col_idx = j
                header_row = i
                break
        if valor_col_idx is not None:
            break

    if valor_col_idx is None:
        return None

    for i in range(header_row + 1, len(df_raw)):
        v = df_raw.iat[i, valor_col_idx]
        f = parse_valor_br(v)
        if f and round(f, 2) != 0.0:
            valores.append(round(f, 2))

    if not valores:
        return None

    s = pd.Series(valores)
    return float(s.mode().iloc[0])


def carregar_boleto_samp(file_like, nome_arquivo: str = "", pasta: str = "") -> dict:
    """
    Lê uma planilha de boleto SAMP (VSP) e identifica sindicato + tipo.
    """
    df_raw = ler_excel_qualquer(file_like, header=None, sheet=0)

    sindicato = identificar_sindicato(nome_arquivo, pasta)
    tipo = identificar_tipo_boleto(df_raw, nome_arquivo)

    # Detecta linha do header e onde começa a tabela
    header_row = None
    for i in range(min(8, len(df_raw))):
        row = [str(x).strip() for x in df_raw.iloc[i].tolist()]
        if "Nome" in row[:2] or "Nome Funcionario" in row or "Nome Funcionário" in row:
            header_row = i
            break

    if header_row is None:
        header_row = 2  # default visto nos arquivos

    headers = [str(x).strip() for x in df_raw.iloc[header_row].tolist()]
    data = df_raw.iloc[header_row + 1:].copy()
    data.columns = headers
    data = data.reset_index(drop=True)

    # Normalizações
    # Possíveis colunas: Nome, CPF, Data Admissão, Sexo, Sindicato,
    # Ultima modificação status, Valor, Status, Serviços
    # OU (controle pagamentos): Nome Funcionario, CPF Funcionário, Data Admissão,
    # Nome Dependente, CPF Dependente, Valor, Serviços
    if "Nome" in data.columns:
        data["Nome"] = data["Nome"].apply(limpar_nome)
    if "Nome Funcionario" in data.columns:
        data["Nome Funcionario"] = data["Nome Funcionario"].apply(limpar_nome)
    if "Nome Dependente" in data.columns:
        data["Nome Dependente"] = data["Nome Dependente"].apply(limpar_nome)
    if "CPF" in data.columns:
        data["CPF"] = data["CPF"].apply(limpar_cpf)
    if "CPF Funcionário" in data.columns:
        data["CPF Funcionário"] = data["CPF Funcionário"].apply(limpar_cpf)
    if "CPF Dependente" in data.columns:
        data["CPF Dependente"] = data["CPF Dependente"].apply(limpar_cpf)
    if "Valor" in data.columns:
        data["Valor_num"] = data["Valor"].apply(parse_valor_br)

    # Remove linhas totalmente vazias
    data = data.dropna(how="all").reset_index(drop=True)

    return {
        "df": data,
        "sindicato": sindicato or "DESCONHECIDO",
        "tipo": tipo,
        "arquivo": nome_arquivo,
    }


def consolidar_samp(boletos: list, dominio_df: pd.DataFrame) -> dict:
    """
    Consolida todos os boletos SAMP (VSP) em uma estrutura única.

    Retorna {
        'boletos': [{sindicato, tipo, df, resumo, valor_total}, ...],
        'resumo_geral': DataFrame consolidado por sindicato/tipo/CCusto
    }
    """
    mapa_ccusto = {}
    if "cpf" in dominio_df.columns and "nome_quebra" in dominio_df.columns:
        for cpf, nq in zip(dominio_df["cpf"], dominio_df["nome_quebra"]):
            if cpf and nq and cpf not in mapa_ccusto:
                mapa_ccusto[cpf] = nq

    resultado_boletos = []
    consolidado_rows = []

    for b in boletos:
        df = b["df"].copy()

        # Determina CPF do funcionário titular
        if "CPF Funcionário" in df.columns:
            cpf_col = "CPF Funcionário"
        elif "CPF" in df.columns:
            cpf_col = "CPF"
        else:
            cpf_col = None

        if cpf_col:
            df["CCUSTO"] = df[cpf_col].map(mapa_ccusto).fillna("")
        else:
            df["CCUSTO"] = ""

        # Resumo por CCusto (ignora zerados)
        df_valid = df[df.get("Valor_num", 0).round(2) != 0.0] if "Valor_num" in df.columns else df.iloc[0:0]
        if not df_valid.empty:
            resumo = (
                df_valid.groupby("CCUSTO", dropna=False, sort=False)["Valor_num"]
                .sum()
                .reset_index()
                .rename(columns={"CCUSTO": "CCusto", "Valor_num": "Valor"})
            )
            resumo = resumo[resumo["CCusto"].astype(str).str.strip() != ""]
            resumo["__cod"] = resumo["CCusto"].apply(codigo_nome_quebra)
            resumo = resumo.sort_values(["__cod", "CCusto"]).drop(columns="__cod").reset_index(drop=True)
        else:
            resumo = pd.DataFrame(columns=["CCusto", "Valor"])

        valor_total = float(round(resumo["Valor"].sum(), 2)) if not resumo.empty else 0.0

        resultado_boletos.append({
            "sindicato": b["sindicato"],
            "tipo": b["tipo"],
            "df": df,
            "resumo": resumo,
            "valor_total": valor_total,
            "arquivo": b.get("arquivo", ""),
        })

        for _, r in resumo.iterrows():
            consolidado_rows.append({
                "Sindicato": b["sindicato"],
                "Tipo": b["tipo"],
                "CCusto": r["CCusto"],
                "Valor": r["Valor"],
            })

    resumo_geral = pd.DataFrame(consolidado_rows)
    return {"boletos": resultado_boletos, "resumo_geral": resumo_geral}


def formatar_valor_br(v: float) -> str:
    try:
        s = f"{float(v):,.2f}"
        return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"


def nome_arquivo_samp(sindicato: str, tipo: str, valor_total: float) -> str:
    tipo_safe = tipo.replace(" - ", " ").replace("/", "_")
    val = formatar_valor_br(valor_total).replace("R$ ", "")
    return f"VSP {sindicato} - {tipo_safe} {val}.xlsx"


def gerar_xlsx_samp_boleto(boleto: dict) -> bytes:
    """Gera o XLSX para um boleto individual."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Boleto"

    df = boleto["df"]
    cols = [c for c in df.columns if not c.startswith("__") and c != "Valor_num"]

    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(df.itertuples(index=False), start=2):
        rec = dict(zip(df.columns, row))
        for j, c in enumerate(cols, start=1):
            ws.cell(row=i, column=j, value=rec.get(c, ""))

    # Resumo
    base = len(cols) + 2
    ws.cell(row=1, column=base, value="CCusto").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=base).fill = PatternFill("solid", fgColor="C00000")
    ws.cell(row=1, column=base + 1, value="Valor").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=base + 1).fill = PatternFill("solid", fgColor="C00000")

    for i, row in enumerate(boleto["resumo"].itertuples(index=False), start=2):
        ws.cell(row=i, column=base, value=row.CCusto)
        cell = ws.cell(row=i, column=base + 1, value=float(row.Valor))
        cell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    last_row = len(boleto["resumo"]) + 2
    ws.cell(row=last_row, column=base, value="TOTAL").font = Font(bold=True)
    tcell = ws.cell(row=last_row, column=base + 1, value=float(boleto["valor_total"]))
    tcell.font = Font(bold=True)
    tcell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    for j in range(1, base + 2):
        col = get_column_letter(j)
        max_len = 0
        for r in range(1, min(ws.max_row + 1, 60)):
            v = ws.cell(row=r, column=j).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
