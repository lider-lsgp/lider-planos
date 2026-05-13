"""
Geração da planilha RELAÇÃO CCUSTO - UNIMED.xlsx final.

Estrutura:
  - Uma ABA por empresa: ATIVA, L COMERCIAL, MULTISSERVIÇOS, VSP.
  - Em cada aba, várias TABELAS lado a lado, uma por contrato:
      ATIVA / L COMERCIAL:
        - Unimed - Saúde (Ambulatorial)
        - Unimed - Saúde (Santas)
        - Unimed - Odonto
      MULTISSERVICOS:
        - Unimed - Saúde (Ambulatorial)
        - Unimed - Odonto
      VSP:
        - Unimed - Saúde (Interior)
        - Unimed - Saúde (Metropolitano)
        - VSP Odonto (consolidado SINDSEG + SINDIVIGILANTES, opcional)
  - Cada tabela: [CCusto | Valor], linhas ordenadas por código do nome_quebra.
  - Linha TOTAL ao final de cada tabela.
"""

import io
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ABA_POR_EMPRESA = {
    "ATIVA": "ATIVA",
    "COMERCIAL": "L COMERCIAL",
    "MULTISSERVICOS": "MULTISSERVIÇOS",
    "VSP": "VSP",
}


def _add_tabela(ws, titulo: str, df_resumo: pd.DataFrame, valor_total: float,
                row_start: int, col_start: int, cor_titulo: str = "1F3864") -> int:
    """
    Escreve uma tabela [CCusto/Valor] em (row_start, col_start).
    Retorna o número de colunas usadas (sempre 2).
    """
    # Linha de título (mesclada)
    ws.cell(row=row_start, column=col_start, value=titulo).font = Font(bold=True, color="FFFFFF", size=11)
    ws.cell(row=row_start, column=col_start).fill = PatternFill("solid", fgColor=cor_titulo)
    ws.cell(row=row_start, column=col_start).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_start, column=col_start + 1).fill = PatternFill("solid", fgColor=cor_titulo)
    ws.merge_cells(start_row=row_start, start_column=col_start,
                   end_row=row_start, end_column=col_start + 1)

    # Header
    h = row_start + 1
    ws.cell(row=h, column=col_start, value="CCusto").font = Font(bold=True)
    ws.cell(row=h, column=col_start + 1, value="Valor").font = Font(bold=True)
    ws.cell(row=h, column=col_start).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(row=h, column=col_start + 1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(row=h, column=col_start).alignment = Alignment(horizontal="center")
    ws.cell(row=h, column=col_start + 1).alignment = Alignment(horizontal="center")

    # Dados
    r = h + 1
    if df_resumo is not None and not df_resumo.empty:
        for _, row in df_resumo.iterrows():
            ws.cell(row=r, column=col_start, value=str(row["CCusto"]))
            c = ws.cell(row=r, column=col_start + 1, value=float(row["Valor"]))
            c.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
            r += 1

    # Total
    ws.cell(row=r, column=col_start, value="TOTAL").font = Font(bold=True)
    tcell = ws.cell(row=r, column=col_start + 1, value=float(valor_total))
    tcell.font = Font(bold=True)
    tcell.fill = PatternFill("solid", fgColor="FFF2CC")
    tcell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    # Bordas finas em toda a tabela
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for rr in range(row_start, r + 1):
        for cc in (col_start, col_start + 1):
            ws.cell(row=rr, column=cc).border = border

    return 2  # colunas ocupadas


def gerar_relacao_ccusto(dados_por_empresa: dict) -> bytes:
    """
    Recebe um dict do tipo:
    {
      'ATIVA': {
          'SAUDE_AMBULATORIAL': {'resumo': df, 'valor_total': float},
          'SAUDE_SANTAS':       {'resumo': df, 'valor_total': float},
          'ODONTO':             {'resumo': df, 'valor_total': float},
      },
      'COMERCIAL': {...},
      'MULTISSERVICOS': {...},
      'VSP': {
          'SAUDE_INTERIOR':       {...},
          'SAUDE_METROPOLITANO':  {...},
          'ODONTO_SAMP':          {'boletos': [...], 'resumo_geral': df}   # opcional
      }
    }
    Retorna bytes do XLSX final.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove sheet default

    # Ordem das abas
    ordem = ["ATIVA", "COMERCIAL", "MULTISSERVICOS", "VSP"]

    for emp in ordem:
        nome_aba = ABA_POR_EMPRESA[emp]
        ws = wb.create_sheet(nome_aba)
        info = dados_por_empresa.get(emp, {}) or {}

        col = 1
        if emp != "VSP":
            tabelas = []
            if info.get("SAUDE_AMBULATORIAL"):
                tabelas.append(("Unimed - Saúde (Ambulatorial)", info["SAUDE_AMBULATORIAL"]))
            if info.get("SAUDE_SANTAS"):
                tabelas.append(("Unimed - Saúde (Santas)", info["SAUDE_SANTAS"]))
            if info.get("ODONTO"):
                tabelas.append(("Unimed - Odonto", info["ODONTO"]))
        else:
            tabelas = []
            if info.get("SAUDE_INTERIOR"):
                tabelas.append(("Unimed - Saúde (Interior)", info["SAUDE_INTERIOR"]))
            if info.get("SAUDE_METROPOLITANO"):
                tabelas.append(("Unimed - Saúde (Metropolitano)", info["SAUDE_METROPOLITANO"]))
            if info.get("ODONTO_SAMP"):
                # Consolidado por sindicato/tipo: 1 tabela por (sindicato, tipo)
                samp = info["ODONTO_SAMP"]
                for b in samp.get("boletos", []):
                    label = f'VSP {b["sindicato"]} - {b["tipo"]}'
                    tabelas.append((label, {"resumo": b["resumo"], "valor_total": b["valor_total"]}))

        if not tabelas:
            ws.cell(row=1, column=1, value=f"(sem dados para {nome_aba})").font = Font(italic=True, color="808080")
            continue

        # Cabeçalho geral da aba
        ws.cell(row=1, column=1, value=f"RELAÇÃO CCUSTO - {nome_aba}").font = Font(bold=True, size=14, color="1F3864")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 * max(len(tabelas), 1) + (max(len(tabelas), 1) - 1))

        # Tabelas lado a lado, começando na linha 3
        row_start = 3
        col = 1
        cores = ["1F3864", "C00000", "548235", "7030A0", "BF8F00", "0070C0"]
        for i, (titulo, payload) in enumerate(tabelas):
            df_res = payload.get("resumo", pd.DataFrame())
            valor = payload.get("valor_total", 0.0)
            cor = cores[i % len(cores)]
            _add_tabela(ws, titulo, df_res, valor, row_start, col, cor_titulo=cor)
            col += 3  # 2 colunas de tabela + 1 de espaço

        # Larguras
        for j in range(1, col):
            letra = get_column_letter(j)
            # Coluna 1 da tabela (CCusto) é mais larga, coluna 2 (Valor) média, espaço = pequena
            if (j - 1) % 3 == 0:
                ws.column_dimensions[letra].width = 35
            elif (j - 1) % 3 == 1:
                ws.column_dimensions[letra].width = 16
            else:
                ws.column_dimensions[letra].width = 3

        ws.freeze_panes = "A3"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
