"""
Tratamento dos relatórios de Saúde da Unimed.

Contratos:
  5957 -> Ambulatorial (ATIVA/COMERCIAL/MULTISSERVICOS)
  6040 -> Santas       (ATIVA/COMERCIAL)
  6217 -> Interior VSP
  5964 -> Metropolitano VSP

Estrutura do relatório (Unimed):
  NRCONTRATO | NOTITULAR | CPFTITULAR | NOUSUARIO | CPF_DEPENDENTE | VLFATURADO [| CDLOTACAO]

Saída esperada do tratamento (em XLSX):
  - Linhas do plano com CCusto puxado da Domínio (PROCV por CPF do titular).
  - Tabela lateral com CCusto único (sem duplicados) + Valor total (SOMASES).
  - Linhas com valor zerado são ignoradas no resumo.
  - Valores negativos (ex.: -8,73 e -15,58) são mantidos no detalhe mas o resumo
    soma tudo (corrigindo o boleto real).
"""

import io
import re
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .utils import (
    ler_excel_qualquer,
    limpar_cpf,
    limpar_nome,
    parse_valor_br,
    codigo_nome_quebra,
)
from .empresas import EMPRESAS, empresa_por_contrato


CONTRATO_TIPO = {
    "5957": ("SAUDE (AMBULATORIAL)", "Unimed - Saúde (Ambulatorial)"),
    "6040": ("SAUDE (SANTAS)",       "Unimed - Saúde (Santas)"),
    "6217": ("SAUDE (INTERIOR)",     "Unimed - Saúde (Interior)"),
    "5964": ("SAUDE (METROPOLITANO)","Unimed - Saúde (Metropolitano)"),
}


def carregar_relatorio_saude(file_like, nome_arquivo: str = "") -> dict:
    """Lê a planilha bruta de Saúde da Unimed e devolve um DataFrame normalizado."""
    df = ler_excel_qualquer(file_like, header=0, sheet=0)
    df.columns = [str(c).strip().upper() for c in df.columns]

    obrig = ["NRCONTRATO", "CPFTITULAR", "VLFATURADO"]
    for c in obrig:
        if c not in df.columns:
            raise ValueError(f"Coluna obrigatória ausente: {c}")

    if "NOTITULAR" in df.columns:
        df["NOTITULAR"] = df["NOTITULAR"].apply(limpar_nome)
    if "NOUSUARIO" in df.columns:
        df["NOUSUARIO"] = df["NOUSUARIO"].apply(limpar_nome)
    df["CPFTITULAR"] = df["CPFTITULAR"].apply(limpar_cpf)
    if "CPF_DEPENDENTE" in df.columns:
        df["CPF_DEPENDENTE"] = df["CPF_DEPENDENTE"].apply(limpar_cpf)
    df["VLFATURADO"] = df["VLFATURADO"].apply(parse_valor_br)
    df["NRCONTRATO"] = df["NRCONTRATO"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

    contrato = df["NRCONTRATO"].dropna().iloc[0] if not df["NRCONTRATO"].dropna().empty else ""
    tipo_curto, tipo_longo = CONTRATO_TIPO.get(contrato, ("SAUDE", "Unimed - Saúde"))

    return {
        "df": df,
        "contrato": contrato,
        "tipo_curto": tipo_curto,
        "tipo_longo": tipo_longo,
        "arquivo": nome_arquivo,
    }


def montar_relatorio_saude(saude: dict, dominio_df: pd.DataFrame, empresa_chave: str) -> dict:
    """
    Cruza relatório de saúde com Domínio por CPF do titular,
    monta DataFrame detalhe + DataFrame resumo (CCusto/Valor).
    Retorna dict com 'detalhe', 'resumo', 'valor_total', 'tipo_curto', 'contrato'.
    """
    df = saude["df"].copy()

    # Mapa CPF -> nome_quebra (CCusto) a partir da Domínio
    mapa_ccusto = {}
    if "cpf" in dominio_df.columns and "nome_quebra" in dominio_df.columns:
        for cpf, nq in zip(dominio_df["cpf"], dominio_df["nome_quebra"]):
            if cpf and nq and cpf not in mapa_ccusto:
                mapa_ccusto[cpf] = nq

    df["CCUSTO"] = df["CPFTITULAR"].map(mapa_ccusto).fillna("")

    # Resumo: por CCusto, somando VLFATURADO (inclui negativos)
    resumo = (
        df.groupby("CCUSTO", dropna=False, sort=False)["VLFATURADO"]
        .sum()
        .reset_index()
        .rename(columns={"CCUSTO": "CCusto", "VLFATURADO": "Valor"})
    )
    # Remove vazios e zerados
    resumo = resumo[resumo["CCusto"].astype(str).str.strip() != ""]
    resumo = resumo[resumo["Valor"].round(2) != 0.0]
    # Ordena por código no início do nome_quebra
    resumo["__cod"] = resumo["CCusto"].apply(codigo_nome_quebra)
    resumo = resumo.sort_values(["__cod", "CCusto"]).drop(columns="__cod").reset_index(drop=True)

    valor_total = float(round(resumo["Valor"].sum(), 2))

    return {
        "detalhe": df,
        "resumo": resumo,
        "valor_total": valor_total,
        "tipo_curto": saude["tipo_curto"],
        "tipo_longo": saude["tipo_longo"],
        "contrato": saude["contrato"],
        "empresa": empresa_chave,
    }


def formatar_valor_br(v: float) -> str:
    """Formata número como 'R$ 1.234,56' estilo brasileiro."""
    try:
        s = f"{float(v):,.2f}"
        return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"


def nome_arquivo_relatorio(empresa_chave: str, tipo_curto: str, valor_total: float) -> str:
    curto = EMPRESAS.get(empresa_chave, {}).get("curto", empresa_chave)
    val = formatar_valor_br(valor_total).replace("R$ ", "")
    return f"RELATORIO - {tipo_curto} - {curto} {val}.xlsx"


def gerar_xlsx_saude(payload: dict) -> bytes:
    """Gera o XLSX final formatado de Saúde."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    df_det = payload["detalhe"]
    df_res = payload["resumo"]

    # ------ Cabeçalho de detalhe ------
    cols_det = ["NRCONTRATO", "NOTITULAR", "CPFTITULAR", "NOUSUARIO",
                "CPF_DEPENDENTE", "VLFATURADO", "CCUSTO"]
    cols_det = [c for c in cols_det if c in df_det.columns]
    for j, c in enumerate(cols_det, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(df_det.itertuples(index=False), start=2):
        rec = dict(zip(df_det.columns, row))
        for j, c in enumerate(cols_det, start=1):
            v = rec.get(c, "")
            cell = ws.cell(row=i, column=j, value=v)
            if c == "VLFATURADO":
                cell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    # ------ Resumo lateral (CCusto / Valor) ------
    # Cola a partir da coluna I (9) com cabeçalho na linha 1
    base = len(cols_det) + 2  # 1 coluna de espaço
    ws.cell(row=1, column=base, value="CCusto").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=base, ).fill = PatternFill("solid", fgColor="C00000")
    ws.cell(row=1, column=base, ).alignment = Alignment(horizontal="center")
    ws.cell(row=1, column=base + 1, value="Valor").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=base + 1).fill = PatternFill("solid", fgColor="C00000")
    ws.cell(row=1, column=base + 1).alignment = Alignment(horizontal="center")

    for i, row in enumerate(df_res.itertuples(index=False), start=2):
        ws.cell(row=i, column=base, value=row.CCusto)
        c = ws.cell(row=i, column=base + 1, value=float(row.Valor))
        c.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    # Linha de TOTAL
    last_row = len(df_res) + 2
    ws.cell(row=last_row, column=base, value="TOTAL").font = Font(bold=True)
    tcell = ws.cell(row=last_row, column=base + 1, value=float(payload["valor_total"]))
    tcell.font = Font(bold=True)
    tcell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    # Larguras automáticas básicas
    for j in range(1, base + 2):
        col = get_column_letter(j)
        max_len = 0
        for r in range(1, min(ws.max_row + 1, 50)):
            v = ws.cell(row=r, column=j).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
