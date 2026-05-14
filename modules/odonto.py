"""
Tratamento do relatório Odonto Unimed (arquivo TXT - delimitado por '#').

Estrutura:
  Linhas 0-6: cabeçalho (Demonstrativo, Contrato, Empresa, header da tabela na L7).
  Linhas a partir de 8 até as 3 últimas: dados.
  3 últimas linhas: Total Empresa / Total Contrato.

Colunas da tabela:
  Código | Beneficiário | Matrícula | CPF | Plano | Tipo | Idade |
  Dependência | Data Limite | Data Inclusão | Data Exclusão | Lotacao |
  Rubrica | Coparticipacao | Outros | Mensalidade | Total Família

Regras:
  - CPF vem com pontos e traços; pode perder zeros à esquerda -> tratar para 11 dígitos.
  - Cruzar CPF do colaborador (Tipo=T) com a Domínio para puxar CCusto.
  - Resumo lateral por CCusto somando 'Total Família' (ou mensalidade, conforme caso).
  - Linhas com valor 0 são ignoradas no resumo.
"""

import io
import re
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import (
    limpar_cpf,
    limpar_nome,
    parse_valor_br,
    codigo_nome_quebra,
    ler_txt_odonto,
)
from .empresas import EMPRESAS, empresa_por_cnpj, empresa_por_nome


COLUNAS_ODONTO = [
    "Código", "Beneficiário", "Matrícula", "CPF", "Plano", "Tipo",
    "Idade", "Dependência", "Data Limite", "Data Inclusão", "Data Exclusão",
    "Lotacao", "Rubrica", "Coparticipacao", "Outros", "Mensalidade", "Total Família",
]


def carregar_relatorio_odonto(file_like, nome_arquivo: str = "") -> dict:
    """Lê e parseia o TXT do Odonto."""
    linhas, empresa_nome, cnpj, fatura = ler_txt_odonto(file_like)

    # A 7ª linha (índice) contém o header. As 3 últimas são totais.
    header_idx = 7
    # Procurar header real
    for i, L in enumerate(linhas[:15]):
        if L.startswith("Código#") or L.startswith("Codigo#") or "Beneficiário" in L:
            header_idx = i
            break

    header = linhas[header_idx].split("#")
    # Remove linhas vazias no fim
    while linhas and not linhas[-1].strip():
        linhas.pop()
    # Últimas 2-3 são totais ("Total Empresa", "Total Contrato")
    fim = len(linhas)
    while fim > header_idx + 1 and ("Total " in linhas[fim - 1] or not linhas[fim - 1].strip()):
        fim -= 1

    rows = []
    for L in linhas[header_idx + 1:fim]:
        if not L.strip():
            continue
        parts = L.split("#")
        # Pad
        while len(parts) < len(header):
            parts.append("")
        rows.append(parts[:len(header)])

    df = pd.DataFrame(rows, columns=header)

    # Renomear (padronizar) colunas mínimas necessárias
    rename = {}
    for c in df.columns:
        lc = c.strip()
        if lc in ("CPF", "Cpf", "cpf"):
            rename[c] = "CPF"
        elif lc.startswith("Benefici"):
            rename[c] = "Beneficiário"
        elif lc == "Tipo":
            rename[c] = "Tipo"
        elif lc.startswith("Mensalid"):
            rename[c] = "Mensalidade"
        elif lc.startswith("Total"):
            rename[c] = "Total Família"
        elif lc == "Outros":
            rename[c] = "Outros"
        elif lc.startswith("Rubrica"):
            rename[c] = "Rubrica"
        elif lc.startswith("Lotac"):
            rename[c] = "Lotacao"
        elif lc.startswith("Data Inclus"):
            rename[c] = "Data Inclusão"
        elif lc.startswith("Data Exclus"):
            rename[c] = "Data Exclusão"
    df = df.rename(columns=rename)

    if "CPF" in df.columns:
        df["CPF"] = df["CPF"].apply(limpar_cpf)
    if "Beneficiário" in df.columns:
        df["Beneficiário"] = df["Beneficiário"].apply(limpar_nome)
    if "Mensalidade" in df.columns:
        df["Mensalidade_num"] = df["Mensalidade"].apply(parse_valor_br)
    if "Outros" in df.columns:
        df["Outros_num"] = df["Outros"].apply(parse_valor_br)
    if "Total Família" in df.columns:
        df["TotalFamilia_num"] = df["Total Família"].apply(parse_valor_br)

    # Empresa detectada
    empresa_chave = empresa_por_cnpj(cnpj) or empresa_por_nome(empresa_nome) or empresa_por_nome(nome_arquivo)

    return {
        "df": df,
        "empresa": empresa_chave or "DESCONHECIDA",
        "empresa_nome": empresa_nome,
        "cnpj": cnpj,
        "fatura": fatura,            # número Fatura (NÃO confundir com valor do boleto)
        "arquivo": nome_arquivo,
    }


def montar_relatorio_odonto(odonto: dict, dominio_df: pd.DataFrame) -> dict:
    """
    Cruza Odonto com Domínio via CPF do TITULAR (Tipo='T') para puxar CCusto.
    Resumo por CCusto soma "Total Família" (valor cheio incluindo dependentes).

    Regra de PJ: titulares que aparecem no Odonto mas NÃO estão no Domínio
    (PJs) vão para o CCusto ADM da empresa (posto código 1).
    """
    df = odonto["df"].copy()

    mapa_ccusto = {}
    if "cpf" in dominio_df.columns and "nome_quebra" in dominio_df.columns:
        for cpf, nq in zip(dominio_df["cpf"], dominio_df["nome_quebra"]):
            if cpf and nq and cpf not in mapa_ccusto:
                mapa_ccusto[cpf] = nq

    # Determina CPF do TITULAR para cada linha. Se a linha é T, é o próprio CPF.
    # Se A/D, herda do último T acima (ordem natural do arquivo já é assim).
    df["__CPF_TITULAR"] = ""
    cpf_atual = ""
    for idx, row in df.iterrows():
        tipo = str(row.get("Tipo", "")).strip().upper()
        cpf = str(row.get("CPF", "")).strip()
        if tipo == "T":
            cpf_atual = cpf
        df.at[idx, "__CPF_TITULAR"] = cpf_atual

    df["CCUSTO"] = df["__CPF_TITULAR"].map(mapa_ccusto).fillna("")

    # Fallback PJ: quem não tem CCusto (não está no Domínio) -> CCusto ADM da empresa
    ccusto_adm = _descobrir_ccusto_adm(dominio_df)
    if ccusto_adm:
        df["__EH_PJ"] = df["CCUSTO"].astype(str).str.strip() == ""
        # Só aplica em linhas que têm CPF de titular preenchido (evita lixo)
        mask_pj = df["__EH_PJ"] & df["__CPF_TITULAR"].astype(str).str.strip().ne("")
        df.loc[mask_pj, "CCUSTO"] = ccusto_adm

    # Resumo: usa Total Família (valor cheio do boleto, incluindo dependentes)
    if "TotalFamilia_num" in df.columns:
        valor_col = "TotalFamilia_num"
    elif "Mensalidade_num" in df.columns:
        valor_col = "Mensalidade_num"
    else:
        valor_col = None

    if valor_col is None:
        resumo = pd.DataFrame(columns=["CCusto", "Valor"])
        valor_total = 0.0
    else:
        resumo = (
            df.groupby("CCUSTO", dropna=False, sort=False)[valor_col]
            .sum()
            .reset_index()
            .rename(columns={"CCUSTO": "CCusto", valor_col: "Valor"})
        )
        resumo = resumo[resumo["CCusto"].astype(str).str.strip() != ""]
        resumo = resumo[resumo["Valor"].round(2) != 0.0]
        resumo["__cod"] = resumo["CCusto"].apply(codigo_nome_quebra)
        resumo = resumo.sort_values(["__cod", "CCusto"]).drop(columns="__cod").reset_index(drop=True)
        valor_total = float(round(resumo["Valor"].sum(), 2))

    return {
        "detalhe": df,
        "resumo": resumo,
        "valor_total": valor_total,
        "empresa": odonto["empresa"],
        "tipo_curto": "ODONTO",
        "tipo_longo": "Unimed - Odonto",
        "fatura": odonto.get("fatura", ""),
    }


def _descobrir_ccusto_adm(dominio_df: pd.DataFrame) -> str:
    """
    Descobre o CCusto ADM (posto código 1) da empresa a partir do Domínio.
    Convencionalmente o código 1 do nome_quebra é a sede/ADM
    (ex.: '1 - LIDER LIMPE LIMPEZA COMERCIAL LTDA').
    """
    if dominio_df is None or dominio_df.empty or "nome_quebra" not in dominio_df.columns:
        return ""
    for nq in dominio_df["nome_quebra"].dropna().astype(str):
        nq_strip = nq.strip()
        if nq_strip.startswith("1 -") or nq_strip.startswith("1 –") or nq_strip.startswith("1-"):
            return nq_strip
    return ""


def formatar_valor_br(v: float) -> str:
    try:
        s = f"{float(v):,.2f}"
        return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"


def nome_arquivo_odonto(empresa_chave: str, valor_total: float) -> str:
    curto = EMPRESAS.get(empresa_chave, {}).get("curto", empresa_chave)
    val = formatar_valor_br(valor_total).replace("R$ ", "")
    return f"RELATORIO - ODONTO - {curto} {val}.xlsx"


def gerar_xlsx_odonto(payload: dict) -> bytes:
    """Gera XLSX final do Odonto."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    df_det = payload["detalhe"]
    df_res = payload["resumo"]

    cols_det = [
        "Código", "Beneficiário", "Matrícula", "CPF", "Plano", "Tipo",
        "Idade", "Dependência", "Data Inclusão", "Data Exclusão",
        "Rubrica", "Coparticipacao", "Outros", "Mensalidade",
        "Total Família", "CCUSTO",
    ]
    cols_det = [c for c in cols_det if c in df_det.columns]

    for j, c in enumerate(cols_det, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(df_det.itertuples(index=False), start=2):
        rec = dict(zip(df_det.columns, row))
        for j, c in enumerate(cols_det, start=1):
            ws.cell(row=i, column=j, value=rec.get(c, ""))

    # Resumo lateral
    base = len(cols_det) + 2
    ws.cell(row=1, column=base, value="CCusto").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=base).fill = PatternFill("solid", fgColor="C00000")
    ws.cell(row=1, column=base + 1, value="Valor").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=base + 1).fill = PatternFill("solid", fgColor="C00000")

    for i, row in enumerate(df_res.itertuples(index=False), start=2):
        ws.cell(row=i, column=base, value=row.CCusto)
        cell = ws.cell(row=i, column=base + 1, value=float(row.Valor))
        cell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    last_row = len(df_res) + 2
    ws.cell(row=last_row, column=base, value="TOTAL").font = Font(bold=True)
    tcell = ws.cell(row=last_row, column=base + 1, value=float(payload["valor_total"]))
    tcell.font = Font(bold=True)
    tcell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    for j in range(1, base + 2):
        col = get_column_letter(j)
        max_len = 0
        for r in range(1, min(ws.max_row + 1, 60)):
            v = ws.cell(row=r, column=j).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
